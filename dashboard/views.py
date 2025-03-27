from django.forms import IntegerField
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from test_logic.models import Test, Result, Question, Option, Product, CompletedTest, CompletedQuestion
from django.http import HttpResponse
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.db.models import Count, Q, F, Value, Case, When, FloatField
from django.db.models.functions import Cast, Coalesce
from django.core.paginator import Paginator
import xlsxwriter
from io import BytesIO
from django.db.models import Avg
from datetime import datetime
from django.contrib import messages
from decimal import Decimal
from accounts.models import User, Region
from .forms import AddBalanceForm, AddStudentForm, ResetTestStatusForm
from django.db import models
from django.core.cache import cache
from django.views.decorators.cache import cache_page
from django.conf import settings

@login_required
def test_list(request):
    tests = Product.objects.all()
    return render(request, 'dashboard/test_list.html', {'tests': tests})

@login_required
def profile(request):
    user = request.user
    return render(request, 'dashboard/profile.html', {'user': user})

@login_required
def test_history(request):
    tests = Test.objects.filter(result__student=request.user).distinct()
    return render(request, 'dashboard/test_history.html', {'tests': tests})

@login_required
def history_detail(request, pk):
    test = get_object_or_404(Test, pk=pk)
    results = Result.objects.filter(student=request.user, test=test)
    
    history = []
    for result in results:
        correct_option = Option.objects.filter(question=result.question, is_correct=True).first()
        history.append({
            'result': result,
            'correct_option': correct_option,
        })
    
    return render(request, 'dashboard/history_detail.html', {'test': test, 'history': history})

@login_required
def add_students(request):
    if request.method == 'POST':
        workbook = load_workbook(request.FILES['document'])
        worksheet = workbook.active

        student_region = request.user.region
        school_region = request.user.school

        for row in worksheet.iter_rows():
            username = row[0].value
            password = row[0].value
            first_name = row[1].value
            last_name = row[2].value
            class_name = row[3].value
                
            user_class_name = username
            user_class_name = ''.join(filter(str.isdigit, user_class_name))

            user_class_password = password
            user_class_password = ''.join(filter(str.isdigit, user_class_password))

            # Try to get the user by username or create a new one
            user, created = User.objects.get_or_create(
                username=user_class_name,
                defaults={
                    'password': user_class_password,
                    'first_name': first_name,
                    'last_name': last_name,
                    'class_name': class_name,
                    'region': student_region,
                    'school': school_region
                }
            )

            # If the user already existed, update their information
            if not created:
                user.first_name = first_name
                user.last_name = last_name
                user.class_name = class_name
                user.region = student_region
                user.school = school_region
                user.save()

        return HttpResponse("Ученики добавлены")
    else:
        return render(request, 'dashboard/addstudent.html')

@login_required
def test_statistics(request):
    # Get filter parameters
    region_id = request.GET.get('region')
    school = request.GET.get('school')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    page = request.GET.get('page', 1)
    
    # Remove cache key and cache operations
    # Base queryset with select_related and prefetch_related
    completed_tests = CompletedTest.objects.select_related(
        'user', 
        'user__region',
        'product'
    )
    
    # Apply filters
    if region_id and region_id != '':
        completed_tests = completed_tests.filter(user__region_id=region_id)
    if school:
        completed_tests = completed_tests.filter(user__school__icontains=school)
    if start_date:
        completed_tests = completed_tests.filter(completed_date__gte=start_date)
    if end_date:
        completed_tests = completed_tests.filter(completed_date__lte=end_date)
    
    # Order by completed date
    completed_tests = completed_tests.order_by('-completed_date')
    
    # Handle Excel export before pagination
    if request.GET.get('export') == 'excel':
        return export_to_excel(completed_tests)
    
    # Count total before pagination for pagination display
    total_count = completed_tests.count()
    
    # Paginate the queryset
    paginator = Paginator(completed_tests, 50)
    page_obj = paginator.get_page(page)
    
    # Now prefetch related data only for the current page
    current_page_tests = page_obj.object_list.prefetch_related(
        'completed_questions',
        'completed_questions__question',
        'completed_questions__selected_option',
        'completed_questions__test'
    )
    
    # Process the test statistics directly
    statistics = []
    for completed_test in current_page_tests:
        # Dictionary to store test-specific statistics
        test_stats = {}
        correct_answers = 0
        wrong_answers = 0
        
        # Get all completed questions for this test in one query
        completed_questions = completed_test.completed_questions.all()
        
        for question in completed_questions:
            # Use the product title or a fallback title if needed
            test_title = completed_test.product.title if hasattr(completed_test.product, 'title') else "Completed Test"
            
            # Initialize test stats if not already present
            if test_title not in test_stats:
                test_stats[test_title] = {
                    'correct': 0,
                    'incorrect': 0,
                    'total': 0
                }
            
            # Check if any selected option is correct
            selected_options = list(question.selected_option.all())
            has_correct = any(option.is_correct for option in selected_options)
            
            # Update statistics
            if has_correct:
                correct_answers += 1
                test_stats[test_title]['correct'] += 1
            else:
                wrong_answers += 1
                test_stats[test_title]['incorrect'] += 1
            
            test_stats[test_title]['total'] += 1
        
        total_questions = correct_answers + wrong_answers
        score_percentage = round((correct_answers / total_questions) * 100, 2) if total_questions > 0 else 0
        
        # Convert test_stats dictionary to a list
        test_statistics = [
            {
                'name': test_name,
                'correct': stats['correct'],
                'incorrect': stats['incorrect'],
                'total': stats['total'],
                'percentage': round((stats['correct'] / stats['total']) * 100, 2) if stats['total'] > 0 else 0
            }
            for test_name, stats in test_stats.items()
        ]
        
        statistics.append({
            'completed_test': completed_test,
            'user': completed_test.user,
            'region': completed_test.user.region,
            'school': completed_test.user.school,
            'completed_date': completed_test.completed_date,
            'correct_answers': correct_answers,
            'wrong_answers': wrong_answers,
            'total_questions': total_questions,
            'score_percentage': score_percentage,
            'test_statistics': test_statistics
        })
    
    # Get all regions for the filter dropdown
    regions = Region.objects.all()

    context = {
        'statistics': statistics,
        'page_obj': page_obj,
        'regions': regions,
        'selected_region': region_id,
        'selected_school': school,
        'start_date': start_date,
        'end_date': end_date,
    }

    return render(request, 'dashboard/test_statistics.html', context)

def export_to_excel(completed_tests):
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output, {'constant_memory': True})
    
    # Create formats
    header_format = workbook.add_format({
        'bold': True,
        'align': 'center',
        'valign': 'vcenter',
        'bg_color': '#f0f0f0'
    })
    
    # Create worksheet with basic structure first
    worksheet = workbook.add_worksheet('Общая статистика')
    
    # Define base headers
    base_headers = [
        'Пользователь', 'Регион', 'Школа', 'Дата завершения',
        'Тест', 'Правильных', 'Неправильных', 'Всего вопросов', 'Результат (%)'
    ]
    
    # Set column widths for better readability
    worksheet.set_column(0, 0, 30)  # User name
    worksheet.set_column(1, 1, 20)  # Region
    worksheet.set_column(2, 2, 20)  # School
    worksheet.set_column(3, 3, 20)  # Date
    worksheet.set_column(4, 4, 30)  # Test name
    worksheet.set_column(5, 8, 15)  # Other columns
    
    # Write headers
    for col, header in enumerate(base_headers):
        worksheet.write(0, col, header, header_format)
    
    # Process and write data in chunks
    row = 1
    
    # Limit the number of records to process to avoid timeout
    max_records = 10000
    record_count = 0
    
    for completed_test in completed_tests.iterator():
        try:
            if record_count >= max_records:
                break
                
            record_count += 1
            
            # Get user info once
            user_name = f"{completed_test.user.first_name} {completed_test.user.last_name}"
            region = str(completed_test.user.region)
            school = completed_test.user.school
            date = completed_test.completed_date.strftime('%Y-%m-%d %H:%M')
            
            # Get all tests for this completed test
            test_ids = completed_test.tests.values_list('id', 'title')
            
            # Create a dictionary to store test statistics
            test_stats = {}
            for test_id, test_title in test_ids:
                test_stats[test_id] = {
                    'title': test_title,
                    'correct': 0,
                    'wrong': 0,
                    'total': 0
                }
            
            # Get completed questions with their test IDs
            from django.db import connection
            with connection.cursor() as cursor:
                # Execute a raw SQL query to get the correct/incorrect counts by test
                cursor.execute("""
                    SELECT 
                        cq.test_id, 
                        COUNT(CASE WHEN o.is_correct THEN 1 END) as correct_count,
                        COUNT(CASE WHEN NOT o.is_correct THEN 1 END) as wrong_count,
                        COUNT(cq.id) as total_count
                    FROM 
                        test_logic_completedquestion cq
                    JOIN 
                        test_logic_completedquestion_selected_option cqso ON cq.id = cqso.completedquestion_id
                    JOIN 
                        test_logic_option o ON cqso.option_id = o.id
                    WHERE 
                        cq.completed_test_id = %s
                    GROUP BY 
                        cq.test_id
                """, [str(completed_test.id)])
                
                results = cursor.fetchall()
            
            # Process the results
            for test_id, correct_count, wrong_count, total_count in results:
                if test_id in test_stats:
                    test_stats[test_id]['correct'] = correct_count
                    test_stats[test_id]['wrong'] = wrong_count
                    test_stats[test_id]['total'] = total_count
            
            # Write one row per test for this user
            for test_id, stats in test_stats.items():
                if stats['total'] > 0:  # Only write rows for tests with questions
                    # Calculate test-specific percentage
                    test_percentage = round((stats['correct'] / stats['total']) * 100, 2) if stats['total'] > 0 else 0
                    
                    # Write data
                    worksheet.write(row, 0, user_name)
                    worksheet.write(row, 1, region)
                    worksheet.write(row, 2, school)
                    worksheet.write(row, 3, date)
                    worksheet.write(row, 4, stats['title'])
                    worksheet.write(row, 5, stats['correct'])
                    worksheet.write(row, 6, stats['wrong'])
                    worksheet.write(row, 7, stats['total'])
                    worksheet.write(row, 8, test_percentage)
                    
                    row += 1
                
        except Exception as e:
            # Log the error but continue processing other records
            print(f"Error processing completed test {completed_test.id}: {str(e)}")
            continue
    
    workbook.close()
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=статистика_тестов.xlsx'
    return response

@login_required
def add_balance(request):
    if not (request.user.is_staff or request.user.is_superuser or request.user.is_principal):
        messages.error(request, "У вас нет прав для доступа к этой странице.")
        return redirect('test_statistics')
    
    form = AddBalanceForm(request.POST or None)
    
    if request.method == 'POST' and form.is_valid():
        filter_type = form.cleaned_data['filter_type']
        amount = form.cleaned_data['amount']
        set_to_zero = form.cleaned_data.get('set_to_zero', False)
        
        try:
            # Use bulk update with F expressions
            if filter_type == 'all':
                if set_to_zero:
                    users_updated = User.objects.filter(is_active=True, balance__gt=0).update(balance=0)
                else:
                    users_updated = User.objects.filter(is_active=True, balance__lt=1500).update(balance=F('balance') + amount)
                
            elif filter_type == 'region':
                region = form.cleaned_data['region']
                if set_to_zero:
                    users_updated = User.objects.filter(region=region, is_active=True, balance__gt=0).update(balance=0)
                else:
                    users_updated = User.objects.filter(region=region, is_active=True, balance__lt=1500).update(balance=F('balance') + amount)
                
            elif filter_type == 'school':
                school = form.cleaned_data['school']
                if set_to_zero:
                    users_updated = User.objects.filter(school__iexact=school, is_active=True, balance__gt=0).update(balance=0)
                else:
                    users_updated = User.objects.filter(school__iexact=school, is_active=True, balance__lt=1500).update(balance=F('balance') + amount)
                
            elif filter_type == 'specific':
                username = form.cleaned_data['username']
                if set_to_zero:
                    users_updated = User.objects.filter(username=username, balance__gt=0).update(balance=0)
                else:
                    users_updated = User.objects.filter(username=username, balance__lt=1500).update(balance=F('balance') + amount)
            
            if set_to_zero:
                messages.success(request, f"Баланс успешно обнулен для {users_updated} пользователя(ей).")
            else:
                messages.success(request, f"Успешно добавлено {amount} на баланс {users_updated} пользователя(ей) с балансом меньше 1500.")
            return redirect('add_balance2')
            
        except Exception as e:
            messages.error(request, f"Произошла ошибка: {str(e)}")

    # Replace cached data fetching with direct database queries
    total_users = User.objects.filter(is_active=True).count()
    low_balance_users = User.objects.filter(is_active=True, balance__lt=1500).count()
    
    # Replace cached region stats with direct database query
    region_stats = list(Region.objects.annotate(
        user_count=Count('user', filter=Q(user__is_active=True)),
        low_balance_count=Count('user', filter=Q(user__is_active=True, user__balance__lt=1500))
    ).values('name', 'user_count', 'low_balance_count'))
    
    # Replace cached school stats with direct database query
    MAX_SCHOOLS = 100  # Limit to prevent performance issues
    
    school_stats = list(User.objects.filter(is_active=True)
        .exclude(Q(school__isnull=True) | Q(school=''))
        .values('school')
        .annotate(
            user_count=Count('id'),
            low_balance_count=Count('id', filter=Q(balance__lt=1500))
        )
        .order_by('-user_count')[:MAX_SCHOOLS])
    
    # Convert to the expected format
    school_stats = [
        {
            'name': school['school'],
            'user_count': school['user_count'],
            'low_balance_count': school['low_balance_count']
        } for school in school_stats
    ]
    
    context = {
        'form': form,
        'total_users': total_users,
        'low_balance_users': low_balance_users,
        'region_stats': region_stats,
        'school_stats': school_stats,
        'show_pagination': True
    }
    
    return render(request, 'dashboard/add_balance.html', context)

@login_required
def question_management(request):
    # Only staff or superusers can manage questions
    if not (request.user.is_staff or request.user.is_superuser):
        messages.error(request, "У вас нет прав для доступа к этой странице.")
        return redirect('test_statistics')
    
    # Get filter parameters
    product_id = request.GET.get('product')
    test_id = request.GET.get('test')
    search_query = request.GET.get('search')
    page = request.GET.get('page', 1)
    
    # Base queryset with prefetch_related to optimize queries
    # Only show questions where question_usage is True
    questions = Question.objects.filter(question_usage=True).select_related(
        'test', 
        'test__product',
        'source_text'
    ).prefetch_related(
        'options'
    ).order_by('test__title', 'text')
    
    # Apply filters
    if product_id:
        questions = questions.filter(test__product_id=product_id)
    
    if test_id:
        questions = questions.filter(test_id=test_id)
    
    if search_query:
        questions = questions.filter(
            Q(text__icontains=search_query) | 
            Q(text2__icontains=search_query) | 
            Q(text3__icontains=search_query) |
            Q(category__icontains=search_query) |
            Q(theme__icontains=search_query)
        )
    
    # Handle question deletion
    if request.method == 'POST' and 'delete_question' in request.POST:
        question_id = request.POST.get('question_id')
        try:
            question = Question.objects.get(id=question_id)
            test_id = question.test.id  # Save test_id for redirect
            question.delete()
            messages.success(request, "Вопрос успешно удален.")
            
            # Redirect to maintain filters
            redirect_url = f'/dashboard/questions/?test={test_id}'
            if product_id:
                redirect_url += f'&product={product_id}'
            if search_query:
                redirect_url += f'&search={search_query}'
            return redirect(redirect_url)
        except Question.DoesNotExist:
            messages.error(request, "Вопрос не найден.")
    
    # Paginate the queryset
    paginator = Paginator(questions, 20)  # 20 questions per page
    page_obj = paginator.get_page(page)
    
    # Get all products and tests for filter dropdowns
    products = Product.objects.all().order_by('title')
    
    # Only get tests related to the selected product if a product is selected
    if product_id:
        tests = Test.objects.filter(product_id=product_id).order_by('title', 'grade')
    else:
        tests = Test.objects.all().order_by('title', 'grade')
    
    context = {
        'page_obj': page_obj,
        'products': products,
        'tests': tests,
        'selected_product': product_id,
        'selected_test': test_id,
        'search_query': search_query,
    }
    
    return render(request, 'dashboard/question_management.html', context)

@login_required
def reset_test_status(request):
    if not (request.user.is_staff or request.user.is_superuser or request.user.is_principal):
        messages.error(request, "У вас нет прав для доступа к этой странице.")
        return redirect('test_statistics')
    
    form = ResetTestStatusForm(request.POST or None)
    
    # Variables to store users that will be affected
    affected_users = []
    preview_mode = True
    
    if request.method == 'POST':
        if 'preview' in request.POST and form.is_valid():
            filter_type = form.cleaned_data['filter_type']
            
            # Use select_related to optimize region fetching in one query
            base_query = User.objects.filter(test_is_started=True).select_related('region')
            
            # Get users based on filter
            if filter_type == 'all':
                # Limit to a reasonable number for preview to avoid memory issues
                affected_users = base_query[:500]
                
            elif filter_type == 'region':
                region = form.cleaned_data['region']
                affected_users = base_query.filter(region=region)[:500]
                
            elif filter_type == 'school':
                school = form.cleaned_data['school']
                affected_users = base_query.filter(school__iexact=school)[:500]
                
            elif filter_type == 'specific':
                username = form.cleaned_data['username']
                try:
                    user = base_query.get(username=username)
                    affected_users = [user]
                except User.DoesNotExist:
                    affected_users = []
            
            # Keep in preview mode
            preview_mode = True
            
        elif 'apply' in request.POST and form.is_valid():
            filter_type = form.cleaned_data['filter_type']
            users_updated = 0
            
            try:
                # Use update() method directly on querysets for bulk operations
                if filter_type == 'all':
                    users_updated = User.objects.filter(test_is_started=True).update(
                        test_is_started=False, 
                        test_start_time=None
                    )
                    
                elif filter_type == 'region':
                    region = form.cleaned_data['region']
                    users_updated = User.objects.filter(
                        region=region, 
                        test_is_started=True
                    ).update(
                        test_is_started=False, 
                        test_start_time=None
                    )
                    
                elif filter_type == 'school':
                    school = form.cleaned_data['school']
                    users_updated = User.objects.filter(
                        school__iexact=school, 
                        test_is_started=True
                    ).update(
                        test_is_started=False, 
                        test_start_time=None
                    )
                    
                elif filter_type == 'specific':
                    username = form.cleaned_data['username']
                    users_updated = User.objects.filter(
                        username=username, 
                        test_is_started=True
                    ).update(
                        test_is_started=False, 
                        test_start_time=None
                    )
                
                messages.success(request, f"Статус теста успешно сброшен для {users_updated} пользователя(ей).")
                return redirect('reset_test_status')
                
            except Exception as e:
                messages.error(request, f"Произошла ошибка: {str(e)}")
            
            # Exit preview mode after applying changes
            preview_mode = False
    
    # Use aggregation for counting to reduce database queries
    from django.db.models import Count, Q
    
    # Get all statistics in a single query
    total_users = User.objects.count()
    started_test_users = User.objects.filter(test_is_started=True).count()
    
    # Get region statistics with a single query using annotation
    regions = Region.objects.annotate(
        user_count=Count('user'),
        started_test_count=Count('user', filter=Q(user__test_is_started=True))
    )
    
    region_stats = [
        {
            'name': region.name,
            'user_count': region.user_count,
            'started_test_count': region.started_test_count
        }
        for region in regions
    ]
    
    # School statistics with a more efficient query
    school_stats = []
    
    # Only process schools with active tests to reduce processing
    schools_with_tests = User.objects.filter(
        test_is_started=True, 
        school__isnull=False
    ).values('school').distinct()
    
    if schools_with_tests.exists():
        # Get all unique schools
        all_schools = User.objects.exclude(
            school__isnull=True
        ).exclude(
            school=''
        ).values_list('school', flat=True).distinct()
        
        # Use a single query with conditional aggregation
        from django.db.models import Sum, Case, When, IntegerField
        
        school_data = User.objects.filter(
            school__in=all_schools
        ).values('school').annotate(
            user_count=Count('id'),
            started_test_count=Sum(
                Case(
                    When(test_is_started=True, then=1),
                    default=0,
                    output_field=IntegerField()
                )
            )
        ).filter(started_test_count__gt=0)  # Only include schools with active tests
        
        for school in school_data:
            school_stats.append({
                'name': school['school'],
                'user_count': school['user_count'],
                'started_test_count': school['started_test_count']
            })
    
    context = {
        'form': form,
        'total_users': total_users,
        'started_test_users': started_test_users,
        'region_stats': region_stats,
        'school_stats': school_stats,
        'affected_users': affected_users,
        'preview_mode': preview_mode
    }
    
    return render(request, 'dashboard/reset_test_status.html', context)

@login_required
def export_by_date(request):
    if not (request.user.is_staff or request.user.is_superuser or request.user.is_principal):
        messages.error(request, "У вас нет прав для доступа к этой странице.")
        return redirect('test_statistics')
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if not start_date or not end_date:
        messages.error(request, "Необходимо указать начальную и конечную даты.")
        return redirect('test_statistics')
    
    # Get all completed tests within the date range
    completed_tests = CompletedTest.objects.select_related(
        'user', 
        'user__region',
        'product'
    ).filter(
        completed_date__gte=start_date,
        completed_date__lte=end_date
    ).order_by('-completed_date')
    
    # Group by date (using date part only)
    from django.db.models.functions import TruncDate
    from django.db.models import Count
    
    # Annotate with just the date part and group
    dates = completed_tests.annotate(date_only=TruncDate('completed_date')).values('date_only').annotate(count=Count('id')).order_by('date_only')
    
    # If no tests, show error
    if not dates:
        messages.error(request, "Нет данных за указанный период.")
        return redirect('test_statistics')
    
    # Create a ZIP file to hold multiple Excel files
    import zipfile
    import io
    from datetime import datetime
    
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'a', zipfile.ZIP_DEFLATED, False) as zip_file:
        for date_group in dates:
            date_str = date_group['date_only'].strftime('%Y-%m-%d')
            
            # Get tests for this date
            date_tests = completed_tests.filter(completed_date__date=date_group['date_only'])
            
            # Create Excel for this date
            excel_buffer = io.BytesIO()
            create_excel_file(date_tests, excel_buffer)
            
            # Add Excel file to ZIP
            zip_file.writestr(f"tests_{date_str}.xlsx", excel_buffer.getvalue())
    
    # Prepare response
    zip_buffer.seek(0)
    response = HttpResponse(
        zip_buffer.read(),
        content_type='application/zip'
    )
    response['Content-Disposition'] = f'attachment; filename=tests_by_date_{start_date}_to_{end_date}.zip'
    
    return response

@login_required
def export_by_school(request):
    if not (request.user.is_staff or request.user.is_superuser or request.user.is_principal):
        messages.error(request, "У вас нет прав для доступа к этой странице.")
        return redirect('test_statistics')
    
    # Get all schools with completed tests
    from django.db.models import Count
    
    schools = CompletedTest.objects.select_related('user').values('user__school').annotate(count=Count('id')).order_by('user__school')
    
    # If no schools, show error
    if not schools:
        messages.error(request, "Нет данных для экспорта.")
        return redirect('test_statistics')
    
    # Create a ZIP file to hold multiple Excel files
    import zipfile
    import io
    
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'a', zipfile.ZIP_DEFLATED, False) as zip_file:
        for school_group in schools:
            school_name = school_group['user__school'] or 'Unknown_School'
            
            # Sanitize filename (remove special characters)
            import re
            safe_school_name = re.sub(r'[^\w\-_\.]', '_', school_name)
            
            # Get tests for this school
            school_tests = CompletedTest.objects.select_related(
                'user', 
                'user__region',
                'product'
            ).filter(user__school=school_name).order_by('-completed_date')
            
            # Create Excel for this school
            excel_buffer = io.BytesIO()
            create_excel_file(school_tests, excel_buffer)
            
            # Add Excel file to ZIP
            zip_file.writestr(f"tests_{safe_school_name}.xlsx", excel_buffer.getvalue())
    
    # Prepare response
    zip_buffer.seek(0)
    response = HttpResponse(
        zip_buffer.read(),
        content_type='application/zip'
    )
    response['Content-Disposition'] = 'attachment; filename=tests_by_school.zip'
    
    return response

def create_excel_file(completed_tests, output_buffer):
    """Helper function to create Excel files with the same format as export_to_excel"""
    
    workbook = xlsxwriter.Workbook(output_buffer, {'constant_memory': True})
    
    # Create formats
    header_format = workbook.add_format({
        'bold': True,
        'align': 'center',
        'valign': 'vcenter',
        'bg_color': '#f0f0f0'
    })
    
    # Create worksheet with basic structure first
    worksheet = workbook.add_worksheet('Общая статистика')
    
    # Define base headers
    base_headers = [
        'Пользователь', 'Регион', 'Школа', 'Дата завершения',
        'Тест', 'Правильных', 'Неправильных', 'Всего вопросов', 'Результат (%)'
    ]
    
    # Set column widths for better readability
    worksheet.set_column(0, 0, 30)  # User name
    worksheet.set_column(1, 1, 20)  # Region
    worksheet.set_column(2, 2, 20)  # School
    worksheet.set_column(3, 3, 20)  # Date
    worksheet.set_column(4, 4, 30)  # Test name
    worksheet.set_column(5, 8, 15)  # Other columns
    
    # Write headers
    for col, header in enumerate(base_headers):
        worksheet.write(0, col, header, header_format)
    
    # Process and write data
    row = 1
    
    # Limit the number of records to process to avoid timeout
    max_records = 10000
    record_count = 0
    
    for completed_test in completed_tests:
        try:
            if record_count >= max_records:
                break
                
            record_count += 1
            
            # Get user info once
            user_name = f"{completed_test.user.first_name} {completed_test.user.last_name}"
            region = str(completed_test.user.region)
            school = completed_test.user.school
            date = completed_test.completed_date.strftime('%Y-%m-%d %H:%M')
            
            # Get all tests for this completed test
            test_ids = completed_test.tests.values_list('id', 'title')
            
            # Create a dictionary to store test statistics
            test_stats = {}
            for test_id, test_title in test_ids:
                test_stats[test_id] = {
                    'title': test_title,
                    'correct': 0,
                    'wrong': 0,
                    'total': 0
                }
            
            # Get completed questions with their test IDs
            from django.db import connection
            with connection.cursor() as cursor:
                # Execute a raw SQL query to get the correct/incorrect counts by test
                cursor.execute("""
                    SELECT 
                        cq.test_id, 
                        COUNT(CASE WHEN o.is_correct THEN 1 END) as correct_count,
                        COUNT(CASE WHEN NOT o.is_correct THEN 1 END) as wrong_count,
                        COUNT(cq.id) as total_count
                    FROM 
                        test_logic_completedquestion cq
                    JOIN 
                        test_logic_completedquestion_selected_option cqso ON cq.id = cqso.completedquestion_id
                    JOIN 
                        test_logic_option o ON cqso.option_id = o.id
                    WHERE 
                        cq.completed_test_id = %s
                    GROUP BY 
                        cq.test_id
                """, [str(completed_test.id)])
                
                results = cursor.fetchall()
            
            # Process the results
            for test_id, correct_count, wrong_count, total_count in results:
                if test_id in test_stats:
                    test_stats[test_id]['correct'] = correct_count
                    test_stats[test_id]['wrong'] = wrong_count
                    test_stats[test_id]['total'] = total_count
            
            # Write one row per test for this user
            for test_id, stats in test_stats.items():
                if stats['total'] > 0:  # Only write rows for tests with questions
                    # Calculate test-specific percentage
                    test_percentage = round((stats['correct'] / stats['total']) * 100, 2) if stats['total'] > 0 else 0
                    
                    # Write data
                    worksheet.write(row, 0, user_name)
                    worksheet.write(row, 1, region)
                    worksheet.write(row, 2, school)
                    worksheet.write(row, 3, date)
                    worksheet.write(row, 4, stats['title'])
                    worksheet.write(row, 5, stats['correct'])
                    worksheet.write(row, 6, stats['wrong'])
                    worksheet.write(row, 7, stats['total'])
                    worksheet.write(row, 8, test_percentage)
                    
                    row += 1
                
        except Exception as e:
            # Log the error but continue processing other records
            print(f"Error processing completed test {completed_test.id}: {str(e)}")
            continue
    
    workbook.close()